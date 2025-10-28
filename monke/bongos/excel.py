"""Excel bongo implementation.

Creates, updates, and deletes test entities via the real Microsoft Graph API.
"""

import asyncio
import io
import time
import uuid
from typing import Any, Dict, List, Optional

import httpx
from monke.bongos.base_bongo import BaseBongo
from monke.generation.excel import generate_workbook_content
from monke.utils.logging import get_logger

# Try to import openpyxl for Excel file creation
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

GRAPH = "https://graph.microsoft.com/v1.0"


class ExcelBongo(BaseBongo):
    """Bongo for Excel that creates test entities for E2E testing.

    Key responsibilities:
    - Create test Excel workbooks in OneDrive
    - Add worksheets with data containing verification tokens
    - Update worksheets to test incremental sync
    - Delete workbooks to test deletion detection
    - Clean up all test data
    """

    connector_type = "excel"

    def __init__(self, credentials: Dict[str, Any], **kwargs):
        super().__init__(credentials)
        self.access_token: str = credentials["access_token"]
        self.entity_count: int = int(kwargs.get("entity_count", 3))
        self.openai_model: str = kwargs.get("openai_model", "gpt-4.1-mini")
        self.rate_limit_delay = float(kwargs.get("rate_limit_delay_ms", 500)) / 1000.0
        self.logger = get_logger("excel_bongo")

        # Track created resources for cleanup
        self._test_workbook_id: Optional[str] = None
        self._test_workbook_name: Optional[str] = None
        self._worksheets: List[Dict[str, Any]] = []
        self._last_req = 0.0

        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel bongo. Install with: pip install openpyxl"
            )

    async def create_entities(self) -> List[Dict[str, Any]]:
        """Create test Excel workbook with worksheets in OneDrive."""
        self.logger.info(
            f"🥁 Creating Excel test workbook with {self.entity_count} worksheets"
        )
        out: List[Dict[str, Any]] = []

        # Generate tokens for each worksheet
        tokens = [uuid.uuid4().hex[:8] for _ in range(self.entity_count)]

        # Generate workbook content
        test_name = f"TestData_{uuid.uuid4().hex[:8]}"
        filename, worksheet_data = await generate_workbook_content(
            self.openai_model, tokens, test_name
        )
        self._test_workbook_name = filename

        self.logger.info(f"📊 Generated {len(worksheet_data)} worksheets")

        # Create Excel workbook file
        workbook_bytes = self._create_excel_file(worksheet_data)

        async with httpx.AsyncClient(base_url=GRAPH, timeout=60) as client:
            # Step 1: Upload Excel file to OneDrive
            await self._pace()
            self.logger.info(f"📤 Uploading Excel file: {filename}")

            upload_url = f"/me/drive/root:/{filename}:/content"
            r = await client.put(
                upload_url,
                headers={
                    "Authorization": f"Bearer {self.access_token}",
                    "Content-Type": (
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                },
                content=workbook_bytes,
            )

            if r.status_code not in (200, 201):
                self.logger.error(f"Upload failed {r.status_code}: {r.text}")
                r.raise_for_status()

            workbook_file = r.json()
            self._test_workbook_id = workbook_file["id"]
            self.logger.info(
                f"✅ Uploaded workbook: {self._test_workbook_id} - {filename}"
            )

            # Wait for Excel to process the file
            # Excel Online needs more time to fully index content for usedRange API
            self.logger.info(
                "⏳ Waiting for Excel Online to process the uploaded file..."
            )
            await asyncio.sleep(10)

            # Step 2: Get worksheet IDs from the workbook
            await self._pace()
            self.logger.info("📋 Fetching worksheet details from workbook")

            worksheets_url = (
                f"/me/drive/items/{self._test_workbook_id}/workbook/worksheets"
            )
            r = await client.get(worksheets_url, headers=self._hdrs())

            if r.status_code != 200:
                self.logger.error(f"Get worksheets failed {r.status_code}: {r.text}")
                r.raise_for_status()

            worksheets = r.json().get("value", [])
            self.logger.info(f"✅ Found {len(worksheets)} worksheets in workbook")

            # Map worksheets to tokens
            for i, (ws, token) in enumerate(zip(worksheets, tokens)):
                ws_id = ws.get("id")
                ws_name = ws.get("name", f"Sheet{i + 1}")

                ent = {
                    "type": "worksheet",
                    "id": ws_id,
                    "workbook_id": self._test_workbook_id,
                    "name": ws_name,
                    "token": token,
                    "expected_content": token,
                }
                out.append(ent)
                self._worksheets.append(ent)
                self.created_entities.append({"id": ws_id, "name": ws_name})
                self.logger.info(
                    f"📄 Worksheet '{ws_name}' created with token: {token}"
                )

        self.logger.info(f"✅ Created workbook with {len(self._worksheets)} worksheets")
        return out

    def _create_excel_file(self, worksheet_data: List[Any]) -> bytes:
        """Create an Excel file with the given worksheet data.

        Args:
            worksheet_data: List of ExcelWorksheetData objects

        Returns:
            Bytes of the Excel file
        """
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for ws_data in worksheet_data:
            ws = wb.create_sheet(title=ws_data.name)

            # Write headers
            for col_idx, header in enumerate(ws_data.headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data rows
            for row_idx, row in enumerate(ws_data.rows, start=2):
                for col_idx, cell_value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=str(cell_value))

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def update_entities(self) -> List[Dict[str, Any]]:
        """Update worksheets by appending new rows with same tokens."""
        if not self._worksheets:
            return []

        self.logger.info(
            f"🥁 Updating {min(2, len(self._worksheets))} Excel worksheets"
        )
        updated = []

        async with httpx.AsyncClient(base_url=GRAPH, timeout=60) as client:
            for ent in self._worksheets[: min(2, len(self._worksheets))]:
                await self._pace()

                # Get current used range to find next empty row
                range_url = (
                    f"/me/drive/items/{self._test_workbook_id}"
                    f"/workbook/worksheets/{ent['id']}/usedRange"
                )
                r = await client.get(range_url, headers=self._hdrs())

                if r.status_code != 200:
                    self.logger.warning(
                        f"Failed to get range: {r.status_code} - {r.text[:200]}"
                    )
                    continue

                range_data = r.json()
                row_count = range_data.get("rowCount", 1)
                next_row = row_count + 1

                # Append new row with token
                new_data = [
                    [
                        f"Updated {ent['token']}",
                        "Update",
                        f"Token: {ent['token']}",
                        "Test",
                    ]
                ]

                # Update specific range
                update_url = (
                    f"/me/drive/items/{self._test_workbook_id}"
                    f"/workbook/worksheets/{ent['id']}/range(address='A{next_row}:D{next_row}')"
                )

                r = await client.patch(
                    update_url,
                    headers=self._hdrs(),
                    json={"values": new_data},
                )

                if r.status_code in (200, 204):
                    updated.append({**ent, "updated": True})
                    self.logger.info(
                        f"📝 Updated worksheet '{ent['name']}' with token: {ent['token']}"
                    )
                else:
                    self.logger.warning(
                        f"Failed to update worksheet: {r.status_code} - {r.text[:200]}"
                    )

                # Brief delay between updates
                await asyncio.sleep(0.5)

        return updated

    async def delete_entities(self) -> List[str]:
        """Delete the entire test workbook."""
        return await self.delete_specific_entities([])

    async def delete_specific_entities(
        self, entities: List[Dict[str, Any]]
    ) -> List[str]:
        """Delete the test workbook (can't delete individual worksheets easily)."""
        if not self._test_workbook_id:
            return []

        self.logger.info(f"🥁 Deleting Excel workbook: {self._test_workbook_name}")
        deleted: List[str] = []

        async with httpx.AsyncClient(base_url=GRAPH, timeout=30) as client:
            try:
                await self._pace()

                # Delete the workbook file from OneDrive
                r = await client.delete(
                    f"/me/drive/items/{self._test_workbook_id}", headers=self._hdrs()
                )

                if r.status_code == 204:
                    # Return worksheet IDs as deleted
                    deleted = [ws["id"] for ws in self._worksheets]
                    self.logger.info(
                        f"✅ Deleted workbook: {self._test_workbook_name} "
                        f"({len(deleted)} worksheets)"
                    )
                    self._worksheets.clear()
                    self._test_workbook_id = None
                else:
                    self.logger.warning(
                        f"Delete failed: {r.status_code} - {r.text[:200]}"
                    )

            except Exception as e:
                self.logger.warning(f"Delete error: {e}")

        return deleted

    async def cleanup(self):
        """Comprehensive cleanup of all test resources."""
        self.logger.info("🧹 Starting comprehensive Excel cleanup")

        cleanup_stats = {
            "workbooks_deleted": 0,
            "errors": 0,
        }

        try:
            async with httpx.AsyncClient(base_url=GRAPH, timeout=30) as client:
                # Delete current test workbook
                if self._test_workbook_id:
                    self.logger.info(
                        f"🗑️ Deleting test workbook: {self._test_workbook_name}"
                    )
                    deleted = await self.delete_specific_entities([])
                    if deleted:
                        cleanup_stats["workbooks_deleted"] += 1

                # Search for and cleanup any orphaned test workbooks
                await self._cleanup_orphaned_workbooks(client, cleanup_stats)

            self.logger.info(
                f"🧹 Cleanup completed: {cleanup_stats['workbooks_deleted']} "
                f"workbooks deleted, {cleanup_stats['errors']} errors"
            )
        except Exception as e:
            self.logger.error(f"❌ Error during comprehensive cleanup: {e}")

    async def _cleanup_orphaned_workbooks(
        self, client: httpx.AsyncClient, stats: Dict[str, Any]
    ):
        """Find and delete orphaned test workbooks from previous runs."""
        try:
            await self._pace()
            r = await client.get("/me/drive/root/children", headers=self._hdrs())

            if r.status_code == 200:
                files = r.json().get("value", [])

                # Find test workbooks
                test_workbooks = [
                    f
                    for f in files
                    if f.get("name", "").startswith("Monke_")
                    and f.get("name", "").endswith(".xlsx")
                ]

                if test_workbooks:
                    self.logger.info(
                        f"🔍 Found {len(test_workbooks)} orphaned test workbooks"
                    )
                    for wb in test_workbooks:
                        try:
                            await self._pace()
                            del_r = await client.delete(
                                f"/me/drive/items/{wb['id']}",
                                headers=self._hdrs(),
                            )
                            if del_r.status_code == 204:
                                stats["workbooks_deleted"] += 1
                                self.logger.info(
                                    f"✅ Deleted orphaned workbook: {wb.get('name', 'Unknown')}"
                                )
                            else:
                                stats["errors"] += 1
                        except Exception as e:
                            stats["errors"] += 1
                            self.logger.warning(
                                f"⚠️ Failed to delete workbook {wb['id']}: {e}"
                            )
        except Exception as e:
            self.logger.warning(f"⚠️ Could not search for orphaned workbooks: {e}")

    def _hdrs(self) -> Dict[str, str]:
        """Get standard headers for Graph API requests."""
        return {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json",
        }

    async def _pace(self):
        """Rate limiting helper."""
        now = time.time()
        if (delta := now - self._last_req) < self.rate_limit_delay:
            await asyncio.sleep(self.rate_limit_delay - delta)
        self._last_req = time.time()
